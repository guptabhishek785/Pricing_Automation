import streamlit as st
import pandas as pd
import numpy as np
import io
import base64
from openpyxl import load_workbook
import os
import time

# Set page title and layout
st.set_page_config(page_title="RSCS Pricing Tool", layout="wide")


# Function to create a download link for a file
def get_download_link(file_path, file_name, text):
    try:
        with open(file_path, "rb") as f:
            data = f.read()
        b64 = base64.b64encode(data).decode()
        href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{file_name}">{text}</a>'
        return href
    except FileNotFoundError:
        return None

# Function to get download link for in-memory file
def get_binary_file_downloader_html(bin_file, file_name, text):
    b64 = base64.b64encode(bin_file).decode()
    href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{file_name}">{text}</a>'
    return href

# Function to process Excel file using the custom logic
def process_excel(uploaded_file, progress_bar):
    # Update progress
    progress_bar.progress(10, text="Reading input file...")
    # time.sleep(0.5)  # Small delay to show progress
    
    # Define your updated column mapping
    columns_mapping = {
        'Proprietary': 1,
        'National Account Manager': 2,
        'Business Analyst': 3,
        'Project #': 4,
        'Add or Change Date': 5,
        'Category': 6,
        'Sub Category': 7,
        'Manufacturer': 12,
        'Vendor Ship City for KINEXO Landed #1': 13,
        'Ship State': 14,
        'Ship Zip': 15,
        'Country of Origin': 16,
        'Contract Holder': 17,
        'Description': 18,
        'Manufacturer Item #': 19,
        'Brand Item#': 20,
        'DC Xref#': 21,
        'GTIN': 22,
        'UPC': 23,
        'Inbound Price Begins': 24,
        'Inbound Price Expires': 25,
        'Vendor Pricing Date': 26,
        'Pack': 27,
        'Size': 28,
        'Ti': 29,
        'Hi': 30,
        'Double-Stacked': 32,
        'Net Wt': 33,
        'Gross Wt': 34,
        'Case Length': 36,
        'Case Width': 37,
        'Case Height': 38,
        'Section': 39,
        'Shelf Life': 40,
        'SHELF LIFE GUARANTEED TO KINEXO (DAYS)': 41,
        'SHELF LIFE GUARANTEED TO DC (DAYS)': 42,
        'DATE ON CASE (MFTR OR EXPIRED OR BEST BY)': 43,
        'EXAMPLE OF DATE': 44,
        'EXPLANATION OF DATE': 45,
        'KINEXO ReD WH': 46,
        'Manf FOB #1': 48,
        'Pallet Charge ($/pallet)': 61,
        'Vendor Delivered $ to KINEXO': 81,
        'Pallet Charge ($/pallet)': 90,
        'Corp Funding %': 99,
        'Outbound Price Begins': 106,
        'OtB Frt $/Cs': 123,
        'Standard Analysis Y/N': 125,
        'National Delievered': 126,
        'DC #': 128,
        'DC Name_x': 129,
        'DC City': 130,
        'DC State': 131,
        'DC Zip': 132,
        'Monthly Case Volume': 134,
        'Lead Time (Days)': 150,
        'Item MOQ': 151,
        'Floor Load': 152,
        # 'Buyer': 155
    }
    
    # Create a BytesIO buffer to store the uploaded file
    uploaded_file_buffer = io.BytesIO(uploaded_file.read())
    
    # Update progress
    progress_bar.progress(20, text="Reading sheets...")
    time.sleep(0.5)
    
    # Read the different sheets from the uploaded file
    try:
        Project_df = pd.read_excel(uploaded_file_buffer, sheet_name="Project Table")
        uploaded_file_buffer.seek(0)
        
        Vendor_df = pd.read_excel(uploaded_file_buffer, sheet_name="Vendor Table")
        uploaded_file_buffer.seek(0)
        
        Item_df = pd.read_excel(uploaded_file_buffer, sheet_name="Item Spec Table")
        uploaded_file_buffer.seek(0)
        
        Demand_df = pd.read_excel(uploaded_file_buffer, sheet_name="Demand Table")
        uploaded_file_buffer.seek(0)
        
        SupplierCapabilities_df = pd.read_excel(uploaded_file_buffer, sheet_name="Vendor Product & Pricing Table")
        uploaded_file_buffer.seek(0)
        
        CustomerDCs_df = pd.read_excel(uploaded_file_buffer, sheet_name="Customer DC Combinations")
        uploaded_file_buffer.seek(0)
        
        Analyst_Inputs_df = pd.read_excel(uploaded_file_buffer, sheet_name="Analysts_Inputs")
        uploaded_file_buffer.seek(0)
        
    except Exception as e:
        raise Exception(f"Error reading sheets from the uploaded file: {e}. Make sure your file has all required sheets: 'Project Table', 'Vendor Table', 'Item Spec Table', 'Demand Table', 'Vendor Product & Pricing Table', 'Customer DC Combinations', and 'Analyst_Inputs_dfs'.")
    
    # Update progress
    progress_bar.progress(40, text="Processing data...")
    time.sleep(0.5)
    
    # Perform data processing with your updated merge logic
    combined_df = (Project_df.merge(Item_df, how='cross')
                   .merge(SupplierCapabilities_df, how='left', on='Brand Item#')
                   .merge(CustomerDCs_df, how='left', on=['Category', 'Sub Category', 'Section', 'CAW Used2', 'KINEXO ReD WH'])
                   .merge(Vendor_df, how='left', on=['Manufacturer', 'Vendor Ship City for KINEXO Landed #1', 'Ship State'])
                   .merge(Demand_df, how='left', on=['DC #', 'Brand Item#']))
    
    # Update progress
    progress_bar.progress(55, text="Filtering and sorting data...")
    time.sleep(0.5)
    
    # Apply data filtering and sorting
    combined_df['Monthly Case Volume'] = combined_df['Monthly Case Volume'].fillna(0)
    combined_df = combined_df[combined_df['Monthly Case Volume'] > 0]
    combined_df = combined_df.sort_values(['Manufacturer', 'Manufacturer Item #', 'Vendor Ship City for KINEXO Landed #1'], ascending=True)
    
    combined_df['Vendor Program'] = Analyst_Inputs_df['Vendor Program'].iloc[0]
    combined_df['KINEXO #'] = Analyst_Inputs_df['KINEXO #'].iloc[0]
    combined_df['Vendor#'] = Analyst_Inputs_df['Vendor#'].iloc[0]
    combined_df['Add or Change Date'] = Analyst_Inputs_df['Add or Change Date'].iloc[0]
    combined_df['Contract Holder'] = Analyst_Inputs_df['Contract Holder'].iloc[0]
    # combined_df['KINEXO Warehouse #1'] = Analyst_Inputs_df['KINEXO Warehouse #1'].iloc[0]
    combined_df['Corp Funding %'] = Analyst_Inputs_df['Corp Funding %'].iloc[0]
    combined_df['Outbound Price Begins'] = Analyst_Inputs_df['Outbound Price Begins'].iloc[0]
    combined_df['Standard Analysis Y/N'] = Analyst_Inputs_df['Standard Analysis Y/N'].iloc[0]
    combined_df['National Delivered'] = Analyst_Inputs_df['National Delivered'].iloc[0]

    # Update progress
    progress_bar.progress(70, text="Loading template...")
    time.sleep(0.5)
    
    # Load the template file
    template_path = "mat.xlsm"
    
    if not os.path.exists(template_path):
        raise Exception(f"Template file 'mat.xlsm' not found. Make sure it's in the same directory as this app.")
    
    # Load the template workbook
    wb = load_workbook(template_path)
    ws = wb.active
    
    # Start row for data entry
    start_row = 4
    
    # Update progress
    progress_bar.progress(80, text="Writing data to template...")
    time.sleep(0.5)
    
    # Write data to the template
    for column_name, excel_col in columns_mapping.items():
        if column_name in combined_df.columns:
            # Extract the column from the DataFrame
            column_to_copy = combined_df[column_name]
            
            # Write the values into the respective column in the Excel sheet
            for idx, value in enumerate(column_to_copy, start=start_row):
                ws.cell(row=idx, column=excel_col, value="NA" if pd.isna(value) else value)
        else:
            st.warning(f"Column '{column_name}' not found in the processed data. Skipping this column.")
    
    # Update progress
    progress_bar.progress(90, text="Saving processed file...")
    time.sleep(0.5)
    
    # Save the workbook to a BytesIO object
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Final progress update
    progress_bar.progress(100, text="Processing complete!")
    time.sleep(0.5)
    
    return output.getvalue()

# Main app layout
def main():
    # Add CSS for better styling
    st.markdown("""
    <style>
    .main-header {
        text-align: center;
        padding: 20px;
    }
    .logo-container {
        display: flex;
        justify-content: center;
        margin-bottom: 30px;
    }
    .button-container {
        display: flex;
        justify-content: center;
        gap: 20px;
        margin-bottom: 30px;
    }
    </style>
    """, unsafe_allow_html=True)
    
    # Header with company logo
    st.markdown('<div class="main-header">', unsafe_allow_html=True)
    st.title("RSCS Pricing Tool")
    st.markdown('</div>', unsafe_allow_html=True)
    
    # Company logo
    st.markdown('<div class="logo-container">', unsafe_allow_html=True)
    
    # Replace with your logo path or use a placeholder
    logo_path = "logo.jpg"  # You'll need to place this file in the same directory
    
    # Check if logo file exists, otherwise use a placeholder
    if os.path.exists(logo_path):
        st.image(logo_path, width=200)
    else:
        st.info("Place your logo.png file in the same directory as this app.")
    
    st.markdown('</div>', unsafe_allow_html=True)
    
    # Create two columns for buttons
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("Step 1: Download Template")
        # Use your existing template file
        template_path = "template.xlsx"
        
        download_link = get_download_link(template_path, "template.xlsx", "Download Template Excel")
        if download_link:
            st.markdown(download_link, unsafe_allow_html=True)
        else:
            st.error(f"Template file not found at: {template_path}")
            st.info("Make sure your template file is named 'template.xlsx' and is in the same directory as this app.")
    
    with col2:
        st.subheader("Step 2: Upload Your Excel File")
        uploaded_file = st.file_uploader("Choose an Excel file", type=["xlsx", "xls"])
        
        if uploaded_file is not None:
            st.success("File successfully uploaded!")
            
            # Create a progress bar
            progress_bar = st.progress(0, text="Ready to process")
            
            # Process button
            if st.button("Process Excel File"):
                try:
                    # Process the file with progress tracking
                    processed_file = process_excel(uploaded_file, progress_bar)
                    
                    # Provide download link for processed file
                    st.success("Processing complete!")
                    st.markdown(
                        get_binary_file_downloader_html(processed_file, "output.xlsx", "Download Processed Excel File"),
                        unsafe_allow_html=True
                    )
                    
                    # Clear the progress bar
                    progress_bar.empty()
                
                except Exception as e:
                    # Clear the progress bar on error
                    progress_bar.empty()
                    st.error(f"Error during processing: {str(e)}")

if __name__ == "__main__":
    main()
