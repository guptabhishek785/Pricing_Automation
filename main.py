import streamlit as st
import pandas as pd
import numpy as np
import io
import base64
from openpyxl import load_workbook
import os
import time

# Set page title and layout
st.set_page_config(page_title="Excel File Processor", layout="wide")

# Function to create a download link for a file
def get_download_link(file_path, file_name, text):
    try:
        with open(file_path, "rb") as f:
            data = f.read()
        b64 = base64.b64encode(data).decode()
        href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{file_name}">{text}</a>'
        return href
    except FileNotFoundError:
        return None

# Function to get download link for in-memory file
def get_binary_file_downloader_html(bin_file, file_name, text):
    b64 = base64.b64encode(bin_file).decode()
    href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{file_name}">{text}</a>'
    return href

# Function to process Excel file using the custom logic
def process_excel(uploaded_file, progress_bar):
    # Update progress
    progress_bar.progress(10, text="Reading input file...")
    time.sleep(0.5)  # Small delay to show progress
    
    # Define your column mapping (same as in your provided code)
    columns_mapping = {
    'Proprietary': 1,
    'National Account Manager': 2,
    'Business Analyst': 3,
    'Project #': 4,
    # 'Add or Change Date'
    'Category': 6,
    'Sub Category': 7,
    'Manufacturer': 12,
    'Vendor Ship City for KINEXO Landed #1': 13,
    'Ship State': 14,
    'Ship Zip': 15,
    'Country of Origin': 16,
    'Description': 18,
    'Manufacturer Item #': 19,
    'Brand Item#': 20,
    'DC Xref#': 21,
    'GTIN': 22,
    'UPC': 23,
    'Inbound Price Begins': 24,
    'Inbound Price Expires': 25,
    'Vendor Pricing Date': 26,
    'Pack': 27,
    'Size': 28,
    'Ti': 29,
    'Hi': 30,
    'Double-Stacked': 32,
    'Net Wt': 33,
    'Gross Wt': 34,
    'Case Length': 36,
    'Case Width': 37,
    'Case Height': 38,
    'Section': 39,
    'Shelf Life': 40,
    'SHELF LIFE GUARANTEED TO KINEXO (DAYS)': 41,
    'SHELF LIFE GUARANTEED TO DC (DAYS)': 42,
    'DATE ON CASE (MFTR OR EXPIRED OR BEST BY)': 43,
    'EXAMPLE OF DATE': 44,
    'EXPLANATION OF DATE': 45,
    # 'KINEXO ReD WH': 46,
    # 'KINEXO Location #1': 47,
    'Manf FOB #1': 48,
    'Pallet Charge ($/pallet)': 61,
    'Vendor Delivered $ to KINEXO': 81,
    'DC #': 128,
    'DC Name_y': 129,
    'DC City': 130,
    'DC State': 131,
    'DC Zip': 132,
    'Monthly Case Volume': 134,
    'Lead Time (Days)': 150,
    'Item MOQ': 151,
    'Floor Load': 152,
    'Buyer': 155
    }
    
    # Create a BytesIO buffer to store the uploaded file
    uploaded_file_buffer = io.BytesIO(uploaded_file.read())
    
    # Update progress
    progress_bar.progress(30, text="Reading sheets...")
    time.sleep(0.5)
    
    # Read the different sheets from the uploaded file
    try:
        Project_df = pd.read_excel(uploaded_file_buffer, sheet_name="Project Table")
        # Reset the buffer position
        uploaded_file_buffer.seek(0)
        
        Vendor_df = pd.read_excel(uploaded_file_buffer, sheet_name="Vendor Table")
        uploaded_file_buffer.seek(0)
        
        Item_df = pd.read_excel(uploaded_file_buffer, sheet_name="Item Spec Table")
        uploaded_file_buffer.seek(0)
        
        Demand_df = pd.read_excel(uploaded_file_buffer, sheet_name="Demand Table")
    except Exception as e:
        raise Exception(f"Error reading sheets from the uploaded file: {e}. Make sure your file has all required sheets: 'Project Table', 'Vendor Table', 'Item Spec Table', and 'Demand Table'.")
    
    # Update progress
    progress_bar.progress(50, text="Processing data...")
    time.sleep(0.5)
    
    # Perform data processing (matching your custom logic)
    combined_df = (Vendor_df.merge(Item_df, how="cross")).merge(Demand_df, how="left", on="Brand Item#")
    combined_df[['National Account Manager', 'Business Analyst', 'Project #', 'Concept', 'Brand Name']] = Project_df[['National Account Manager', 'Business Analyst', 'Project #', 'Category', 'Sub Category']].iloc[0]
    
    # Update progress
    progress_bar.progress(70, text="Loading template...")
    time.sleep(0.5)
    
    # Load the template file
    template_path = "mat.xlsm"
    
    if not os.path.exists(template_path):
        raise Exception(f"Template file 'mat.xlsm' not found. Make sure it's in the same directory as this app.")
    
    # Load the template workbook
    wb = load_workbook(template_path)
    ws = wb.active
    
    # Start row for data entry
    start_row = 4
    
    # Update progress
    progress_bar.progress(80, text="Writing data to template...")
    time.sleep(0.5)
    
    # Write data to the template
    for column_name, excel_col in columns_mapping.items():
        if column_name in combined_df.columns:
            # Extract the column from the DataFrame
            column_to_copy = combined_df[column_name]
            
            # Write the values into the respective column in the Excel sheet
            for idx, value in enumerate(column_to_copy, start=start_row):
                ws.cell(row=idx, column=excel_col, value=value)
        else:
            st.warning(f"Column '{column_name}' not found in the processed data. Skipping this column.")
    
    # Update progress
    progress_bar.progress(90, text="Saving processed file...")
    time.sleep(0.5)
    
    # Save the workbook to a BytesIO object
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Final progress update
    progress_bar.progress(100, text="Processing complete!")
    time.sleep(0.5)
    
    return output.getvalue()

# Main app layout
def main():
    # Add CSS for better styling
    st.markdown("""
    <style>
    .main-header {
        text-align: center;
        padding: 20px;
    }
    .logo-container {
        display: flex;
        justify-content: center;
        margin-bottom: 30px;
    }
    .button-container {
        display: flex;
        justify-content: center;
        gap: 20px;
        margin-bottom: 30px;
    }
    </style>
    """, unsafe_allow_html=True)
    
    # Header with company logo
    st.markdown('<div class="main-header">', unsafe_allow_html=True)
    st.title("Excel Data Processor")
    st.markdown('</div>', unsafe_allow_html=True)
    
    # Company logo
    st.markdown('<div class="logo-container">', unsafe_allow_html=True)
    
    # Replace with your logo path or use a placeholder
    logo_path = "logo.jpg"  # You'll need to place this file in the same directory
    
    # Check if logo file exists, otherwise use a placeholder
    if os.path.exists(logo_path):
        st.image(logo_path, width=200)
    else:
        st.info("Place your logo.png/ logo.jpg file in the same directory as this app.")
    
    st.markdown('</div>', unsafe_allow_html=True)
    
    # Create two columns for buttons
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("Step 1: Download Template")
        # Use your existing template file
        template_path = "template.xlsx"
        
        download_link = get_download_link(template_path, "template.xlsx", "Download Template Excel")
        if download_link:
            st.markdown(download_link, unsafe_allow_html=True)
        else:
            st.error(f"Template file not found at: {template_path}")
            st.info("Make sure your template file is named 'template.xlsx' and is in the same directory as this app.")
    
    with col2:
        st.subheader("Step 2: Upload Your Excel File")
        uploaded_file = st.file_uploader("Choose an Excel file", type=["xlsx", "xls"])
        
        if uploaded_file is not None:
            st.success("File successfully uploaded!")
            
            # Create a progress bar
            progress_bar = st.progress(0, text="Ready to process")
            
            # Process button
            if st.button("Process Excel File"):
                try:
                    # Process the file with progress tracking
                    processed_file = process_excel(uploaded_file, progress_bar)
                    
                    # Provide download link for processed file
                    st.success("Processing complete!")
                    st.markdown(
                        get_binary_file_downloader_html(processed_file, "output.xlsx", "Download Processed Excel File"),
                        unsafe_allow_html=True
                    )
                    
                    # Clear the progress bar
                    progress_bar.empty()
                
                except Exception as e:
                    # Clear the progress bar on error
                    progress_bar.empty()
                    st.error(f"Error during processing: {str(e)}")

if __name__ == "__main__":
    main()